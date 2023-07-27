import os
import yaml
import openpyxl
from jnius import autoclass
from kivy.utils import platform
from datetime import datetime
from screeninfo import get_monitors



def get_project_dir():
    return os.path.abspath(os.path.dirname(os.path.dirname(__file__)))



def read_yaml(path, filename):
    with open(os.path.join(path, f"{filename}.yaml"), 'r') as f:
        file = yaml.safe_load(f)
    return file



PATH_TO_CONFIGS = os.path.join(get_project_dir(), 'configs')
SCR_RES = read_yaml(path=PATH_TO_CONFIGS, filename='screen_resolution')
CONFIGS_MODEL = read_yaml(path=PATH_TO_CONFIGS, filename='configs_model')



def get_window_size():
    if platform == 'android':
        PythonActivity = autoclass("org.kivy.android.PythonActivity")
        activity = PythonActivity.mActivity
        displayMetrics = activity.getContext().getResources().getDisplayMetrics()
        width = displayMetrics.widthPixels
        height = displayMetrics.heightPixels
    else:
        width = get_monitors()[0].width
        height = get_monitors()[0].height
    return width, height



def update_window_size():
    width, height = get_window_size()
    SCR_RES['width'] = width
    SCR_RES['height'] = height
    with open(os.path.join(PATH_TO_CONFIGS, f"screen_resolution.yaml"), 'w') as f:
        yaml.dump(SCR_RES, f)



def time_converter(dt):
    hours = int(dt // 3600)
    dt = dt % 3600
    minutes = int(dt // 60)
    dt = dt % 60
    seconds = int(dt)
    if len(str(hours)) == 1: hours = f'0{hours}'
    if len(str(minutes)) == 1: minutes = f'0{minutes}'
    if len(str(seconds)) == 1: seconds = f'0{seconds}'
    return f'{hours}:{minutes}:{seconds}'



def time_diff(t1, t2):
    FMT = '%H:%M:%S'
    dt = (datetime.strptime(t2, FMT) - datetime.strptime(t1, FMT)).total_seconds()
    hours = int(dt // 3600)
    dt = dt % 3600
    minutes = int(dt // 60)
    dt = dt % 60
    seconds = int(dt)
    if len(str(hours)) == 1: hours = f'0{hours}'
    if len(str(minutes)) == 1: minutes = f'0{minutes}'
    if len(str(seconds)) == 1: seconds = f'0{seconds}'
    return f'{hours}:{minutes}:{seconds}'



def save(data, configs, start_time):
    converter = dict([(key, it) for (it, key) in enumerate(CONFIGS_MODEL.keys())])
    converter2 = dict([(val, key) for (key, val) in CONFIGS_MODEL[configs['task_type']]['conf3']['actual_values'].items()])
    path = 'logs'
    n_true = sum([ans1 == ans2 for (ans1, ans2) in zip(data['Answer'], data['YourAnswer'])])
    filename1 = f"{start_time.strftime('%Y:%m:%d_%H:%M:%S')}_{converter[configs['task_type']]}_{n_true}"
    filename2 = f"_{configs['n_tasks']}_{configs['max_number']}_{converter2[configs['conf3']]}.xlsx"
    filename = filename1 + filename2
    fields = ['№', 'Time', 'Answer', 'YourAnswer', 'Task']
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    for it, val in enumerate(fields):
        c = my_sheet.cell(row=1, column=it+1)
        c.value = val
    for ix in range(len(data['YourAnswer'])):
        for iy, key in enumerate(fields):
            c = my_sheet.cell(row=ix+2, column=iy + 1)
            c.value = data[key][ix]
    my_wb.save(os.path.join(path, filename))



def read(filename):
    path = 'logs'
    fields = ['№', 'Time', 'Answer', 'YourAnswer', 'Task']
    my_wb = openpyxl.load_workbook(os.path.join(path, filename))
    my_sheet = my_wb.active
    data = dict(zip(fields, [[] for _ in range(len(fields))]))
    for ix in range(my_sheet.max_row-1):
        for iy, key in enumerate(fields):
            c = my_sheet.cell(row=ix+2, column=iy + 1)
            data[key].append(c.value)
    return data



